"""
管理员API
"""

from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_

from app import db
from app.models.user import User
from app.models.alumni_profile import AlumniProfile
from app.models.visit_application import VisitApplication
from app.models.visit_record import VisitRecord
from app.models.student_exit_application import StudentExitApplication
from app.models.organization import Organization

admin_bp = Blueprint('admin', __name__)

@admin_bp.before_request
def admin_required():
    """管理员权限检查"""
    # OPTIONS请求用于CORS预检，不需要身份验证
    if request.method == 'OPTIONS':
        return

    # 检查JWT token
    try:
        from flask_jwt_extended import verify_jwt_in_request
        verify_jwt_in_request()
    except Exception:
        return jsonify({'error': '需要有效的认证token'}), 401

    # 检查管理员权限
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)

    if not current_user or current_user.user_type != 'admin':
        return jsonify({'error': '需要管理员权限'}), 403

@admin_bp.route('/dashboard', methods=['GET'])
def get_dashboard():
    """获取仪表板数据"""
    try:
        # 基础统计数据
        total_users = User.query.count()
        total_alumni = User.query.filter_by(user_type='alumni').count()
        total_teachers = User.query.filter_by(user_type='teacher').count()
        total_security = User.query.filter_by(user_type='security').count()
        total_students = User.query.filter_by(user_type='student').count()
        total_parents = User.query.filter_by(user_type='parent').count()

        # 访问相关统计
        today = datetime.utcnow().date()
        today_visits = VisitRecord.query.filter(
            func.date(VisitRecord.entry_time) == today
        ).count()

        total_visits = VisitRecord.query.count()
        total_applications = VisitApplication.query.count()

        # 学生出校相关统计
        total_students_exits = StudentExitApplication.query.count()
        today_student_exits = StudentExitApplication.query.filter(
            func.date(StudentExitApplication.created_at) == today
        ).count()
        approved_student_exits = StudentExitApplication.query.filter_by(application_status='approved').count()
        pending_student_exits = StudentExitApplication.query.filter_by(application_status='pending').count()

        # 待审核事项
        pending_alumni = AlumniProfile.query.filter_by(approval_status='pending').count()
        pending_visits = VisitApplication.query.filter_by(application_status='pending').count()
        pending_parent_approvals = StudentExitApplication.query.filter_by(parent_approval_status='pending').count()
        pending_teacher_approvals = StudentExitApplication.query.filter_by(teacher_approval_status='pending').count()

        # 模拟趋势数据（避免复杂查询导致的错误）
        from datetime import date, timedelta
        visit_trend = []
        student_exit_trend = []
        for i in range(7):
            d = date.today() - timedelta(days=6-i)
            visit_count = VisitRecord.query.filter(
                func.date(VisitRecord.entry_time) == d
            ).count()
            exit_count = StudentExitApplication.query.filter(
                func.date(StudentExitApplication.created_at) == d
            ).count()
            visit_trend.append({'date': d.isoformat(), 'count': visit_count})
            student_exit_trend.append({'date': d.isoformat(), 'count': exit_count})

        # 访问目的统计
        purpose_stats = db.session.query(
            VisitApplication.visit_purpose,
            func.count(VisitApplication.id).label('count')
        ).filter(
            VisitApplication.application_status.in_(['approved', 'completed'])
        ).group_by(VisitApplication.visit_purpose).limit(10).all()

        # 学生出校原因统计
        exit_reason_stats = db.session.query(
            StudentExitApplication.exit_reason,
            func.count(StudentExitApplication.id).label('count')
        ).filter(
            StudentExitApplication.application_status.in_(['approved', 'completed'])
        ).group_by(StudentExitApplication.exit_reason).limit(10).all()

        # 审批状态统计
        parent_approval_stats = db.session.query(
            StudentExitApplication.parent_approval_status,
            func.count(StudentExitApplication.id).label('count')
        ).group_by(StudentExitApplication.parent_approval_status).all()

        teacher_approval_stats = db.session.query(
            StudentExitApplication.teacher_approval_status,
            func.count(StudentExitApplication.id).label('count')
        ).group_by(StudentExitApplication.teacher_approval_status).all()

        return jsonify({
            'statistics': {
                'total_users': total_users,
                'total_alumni': total_alumni,
                'total_teachers': total_teachers,
                'total_security': total_security,
                'total_students': total_students,
                'total_parents': total_parents,
                'today_visits': today_visits,
                'total_visits': total_visits,
                'total_applications': total_applications,
                'total_students_exits': total_students_exits,
                'today_student_exits': today_student_exits,
                'approved_student_exits': approved_student_exits,
                'pending_student_exits': pending_student_exits,
                'pending_alumni': pending_alumni,
                'pending_visits': pending_visits,
                'pending_parent_approvals': pending_parent_approvals,
                'pending_teacher_approvals': pending_teacher_approvals
            },
            'visit_trend': visit_trend,
            'student_exit_trend': student_exit_trend,
            'purpose_stats': [{'purpose': stat.visit_purpose, 'count': stat.count} for stat in purpose_stats],
            'exit_reason_stats': [{'reason': stat.exit_reason, 'count': stat.count} for stat in exit_reason_stats],
            'parent_approval_stats': [{'status': stat.parent_approval_status, 'count': stat.count} for stat in parent_approval_stats],
            'teacher_approval_stats': [{'status': stat.teacher_approval_status, 'count': stat.count} for stat in teacher_approval_stats]
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取仪表板数据失败: {str(e)}")
        return jsonify({'error': '获取仪表板数据失败', 'details': str(e)}), 500

@admin_bp.route('/users', methods=['POST'])
def create_user():
    """创建新用户"""
    try:
        data = request.get_json()

        # 必填字段验证
        required_fields = ['username', 'password', 'real_name', 'user_type', 'email']
        for field in required_fields:
            if not data.get(field) or str(data.get(field)).strip() == '':
                return jsonify({'error': f'{field} 不能为空'}), 400

        # 用户类型验证 - 支持多用户类型（逗号分隔）
        valid_user_types = ['admin', 'teacher', 'student', 'parent', 'alumni', 'security']

        # 处理多用户类型
        user_types = [ut.strip() for ut in data['user_type'].split(',') if ut.strip()]

        # 验证每个用户类型都有效
        for ut in user_types:
            if ut not in valid_user_types:
                return jsonify({'error': f'用户类型无效，必须是: {", ".join(valid_user_types)}'}), 400

        # 重新组合用户类型字符串
        data['user_type'] = ','.join(user_types)

        # 检查用户名是否已存在
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': '用户名已存在'}), 400

        # 检查邮箱是否已存在
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': '邮箱已存在'}), 400

        # 检查手机号是否已存在（如果提供）
        if data.get('phone') and User.query.filter_by(phone=data['phone']).first():
            return jsonify({'error': '手机号已存在'}), 400

        # 创建新用户
        import uuid
        new_user = User(
            uuid=str(uuid.uuid4()),
            username=data['username'],
            real_name=data['real_name'],
            user_type=data['user_type'],
            email=data['email'],
            phone=data.get('phone', ''),
            student_id=data.get('student_id', ''),
            employee_id=data.get('employee_id', ''),
            class_id=data.get('class_id', ''),
            grade=data.get('grade', ''),
            status='active'
        )

        # 设置密码
        new_user.set_password(data['password'])

        # 设置特定类型的字段
        if 'teacher' in data['user_type'] and data.get('is_class_teacher'):
            new_user.is_class_teacher = True
        elif 'student' in data['user_type'] and data.get('student_parent_id'):
            new_user.student_parent_id = data['student_parent_id']

        # 自动设置可拜访权限 - 只要是教师就自动设为可拜访
        if 'teacher' in data['user_type']:
            new_user.is_visitable = True

        # 保存用户
        db.session.add(new_user)
        db.session.commit()

        current_app.logger.info(f"管理员创建用户成功: {new_user.username} ({new_user.user_type})")

        return jsonify({
            'message': '用户创建成功',
            'user': new_user.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"创建用户失败: {str(e)}")
        return jsonify({'error': '创建用户失败', 'details': str(e)}), 500

@admin_bp.route('/users', methods=['GET'])
def get_users():
    """获取用户列表"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        user_type = request.args.get('user_type', '')
        status = request.args.get('status', '')
        search = request.args.get('search', '')

        query = User.query

        if user_type:
            query = query.filter_by(user_type=user_type)

        if status:
            query = query.filter_by(status=status)

        if search:
            query = query.filter(
                User.real_name.contains(search) |
                User.username.contains(search) |
                User.phone.contains(search) |
                User.email.contains(search)
            )

        # 关联校友档案搜索
        if search:
            query = query.outerjoin(AlumniProfile).filter(
                AlumniProfile.student_id.contains(search) |
                AlumniProfile.class_name.contains(search) |
                AlumniProfile.division.contains(search)
            )

        query = query.order_by(User.created_at.desc())

        pagination = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )

        users_list = []
        for user in pagination.items:
            user_data = user.to_dict()
            if user.alumni_profile:
                user_data['alumni_profile'] = user.alumni_profile.to_dict()
            users_list.append(user_data)

        return jsonify({
            'users': users_list,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取用户列表失败: {str(e)}")
        return jsonify({'error': '获取用户列表失败'}), 500

@admin_bp.route('/users/<int:user_id>/status', methods=['PUT'])
def update_user_status(user_id):
    """更新用户状态"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': '用户不存在'}), 404

        data = request.get_json()
        status = data.get('status')

        if status not in ['active', 'inactive']:
            return jsonify({'error': '状态值无效'}), 400

        user.status = status
        user.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'message': '用户状态更新成功',
            'user': user.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新用户状态失败: {str(e)}")
        return jsonify({'error': '更新用户状态失败'}), 500

@admin_bp.route('/alumni-approve', methods=['GET'])
def get_alumni_for_approval():
    """获取校友审核列表 - 支持状态筛选"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        status = request.args.get('status', 'pending')  # pending, approved, rejected, all
        search = request.args.get('search', '')
        division = request.args.get('division', '')
        graduation_year = request.args.get('graduation_year', '')

        # 构建查询，明确指定join关系
        query = AlumniProfile.query.join(User, AlumniProfile.user_id == User.id)

        # 状态筛选
        if status != 'all':
            query = query.filter(AlumniProfile.approval_status == status)

        # 搜索筛选
        if search:
            query = query.filter(
                db.or_(
                    User.real_name.contains(search),
                    User.username.contains(search),
                    AlumniProfile.student_id.contains(search),
                    AlumniProfile.class_name.contains(search)
                )
            )

        # 学部筛选
        if division:
            query = query.filter(AlumniProfile.division == division)

        # 年级筛选
        if graduation_year:
            query = query.filter(AlumniProfile.graduation_year == graduation_year)

        # 排序：待审核优先，然后按创建时间倒序
        query = query.order_by(
            db.case(
                (AlumniProfile.approval_status == 'pending', 0),
                (AlumniProfile.approval_status == 'rejected', 2),
                (AlumniProfile.approval_status == 'approved', 3),
                else_=1
            ),
            AlumniProfile.created_at.desc()
        )

        pagination = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )

        alumni_list = []
        for profile in pagination.items:
            alumni_data = profile.to_dict()
            alumni_data['user'] = profile.user.to_dict()
            # 添加审核历史信息
            if profile.approved_by:
                approver = User.query.get(profile.approved_by)
                alumni_data['approver_name'] = approver.real_name if approver else '未知'
            else:
                alumni_data['approver_name'] = None
            alumni_list.append(alumni_data)

        # 获取统计数据
        stats = {
            'pending': AlumniProfile.query.filter_by(approval_status='pending').count(),
            'approved': AlumniProfile.query.filter_by(approval_status='approved').count(),
            'rejected': AlumniProfile.query.filter_by(approval_status='rejected').count(),
            'total': AlumniProfile.query.count()
        }

        # 获取筛选选项
        divisions = db.session.query(AlumniProfile.division).distinct().all()
        graduation_years = db.session.query(AlumniProfile.graduation_year).distinct().order_by(AlumniProfile.graduation_year.desc()).all()

        return jsonify({
            'alumni': alumni_list,
            'statistics': stats,
            'filters': {
                'divisions': [div[0] for div in divisions if div[0]],
                'graduation_years': [year[0] for year in graduation_years]
            },
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取校友审核列表失败: {str(e)}")
        return jsonify({'error': '获取校友审核列表失败', 'details': str(e)}), 500

@admin_bp.route('/alumni/<int:profile_id>/approve', methods=['POST'])
def approve_alumni(profile_id):
    """审核校友"""
    try:
        current_user_id = get_jwt_identity()

        profile = AlumniProfile.query.get(profile_id)
        if not profile:
            return jsonify({'error': '校友档案不存在'}), 404

        if profile.approval_status != 'pending':
            return jsonify({'error': '该档案已经处理过了'}), 400

        data = request.get_json()
        approve = data.get('approve', True)
        note = data.get('note', '')

        profile.approval_status = 'approved' if approve else 'rejected'
        profile.approved_by = current_user_id
        profile.approval_time = datetime.utcnow()
        profile.approval_note = note

        # 如果审核通过，激活用户账户
        user = profile.user
        if approve and user.status == 'pending':
            user.status = 'active'

        db.session.commit()

        return jsonify({
            'message': f'校友档案{"通过" if approve else "拒绝"}审核',
            'alumni': profile.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"审核校友档案失败: {str(e)}")
        return jsonify({'error': '审核校友档案失败'}), 500

@admin_bp.route('/alumni/<int:profile_id>/reapprove', methods=['POST'])
def reapprove_alumni(profile_id):
    """重新审核校友"""
    try:
        current_user_id = get_jwt_identity()

        profile = AlumniProfile.query.get(profile_id)
        if not profile:
            return jsonify({'error': '校友档案不存在'}), 404

        # 只允许重新审核已拒绝的校友
        if profile.approval_status != 'rejected':
            return jsonify({'error': '只能重新审核已拒绝的校友档案'}), 400

        # 重置审核状态
        profile.approval_status = 'pending'
        profile.approved_by = None
        profile.approval_time = None
        profile.approval_note = ''

        db.session.commit()

        return jsonify({
            'message': '校友档案已重新提交审核',
            'alumni': profile.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"重新审核校友档案失败: {str(e)}")
        return jsonify({'error': '重新审核校友档案失败'}), 500

@admin_bp.route('/batch-approve', methods=['POST'])
def batch_approve():
    """批量授权访问"""
    try:
        current_user_id = get_jwt_identity()

        data = request.get_json()
        batch_type = data.get('type')  # 'class', 'year', 'division'
        target_value = data.get('target_value')
        visit_date = data.get('visit_date')
        time_start = data.get('time_start')
        time_end = data.get('time_end')
        visit_purpose = data.get('visit_purpose')

        if not all([batch_type, target_value, visit_date, time_start, time_end, visit_purpose]):
            return jsonify({'error': '请提供完整的批量授权信息'}), 400

        # 查找符合条件的校友
        query = AlumniProfile.query.filter_by(approval_status='approved')
        if batch_type == 'class':
            query = query.filter_by(class_name=target_value)
        elif batch_type == 'year':
            query = query.filter_by(graduation_year=int(target_value))
        elif batch_type == 'division':
            query = query.filter_by(division=target_value)
        else:
            return jsonify({'error': '批量授权类型无效'}), 400

        alumni_profiles = query.all()

        if not alumni_profiles:
            return jsonify({'error': '未找到符合条件的校友'}), 404

        # 解析时间
        visit_date = datetime.strptime(visit_date, '%Y-%m-%d').date()
        time_start = datetime.strptime(time_start, '%H:%M').time()
        time_end = datetime.strptime(time_end, '%H:%M').time()

        # 创建访问申请
        created_applications = []
        for profile in alumni_profiles:
            application = VisitApplication(
                applicant_id=profile.user_id,
                visit_date=visit_date,
                visit_time_start=time_start,
                visit_time_end=time_end,
                visit_purpose=f"批量授权: {visit_purpose}",
                application_status='approved',
                approved_by=current_user_id,
                approval_time=datetime.utcnow(),
                approval_note=f"批量授权 - {batch_type}: {target_value}"
            )
            db.session.add(application)
            created_applications.append(application)

        db.session.commit()

        return jsonify({
            'message': f'成功为 {len(created_applications)} 位校友创建访问授权',
            'count': len(created_applications)
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量授权失败: {str(e)}")
        return jsonify({'error': '批量授权失败'}), 500

@admin_bp.route('/statistics', methods=['GET'])
def get_statistics():
    """获取详细统计数据"""
    try:
        # 时间范围参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        group_by = request.args.get('group_by', 'day')  # day, week, month
        statistic_type = request.args.get('type', 'overview')  # overview, visits, applications, people

        # 基础统计
        total_visits = VisitRecord.query.count()
        total_applications = VisitApplication.query.count()
        total_users = User.query.count()
        total_alumni = User.query.filter_by(user_type='alumni').count()

        # 导入学生出校申请模型
        try:
            from app.models.student_exit_application import StudentExitApplication
            total_student_exit_applications = StudentExitApplication.query.count()
        except ImportError:
            total_student_exit_applications = 0

        if statistic_type == 'overview':
            # 概览统计
            today = datetime.utcnow().date()

            # 时间段统计
            today_visits = VisitRecord.query.filter(
                func.date(VisitRecord.entry_time) == today
            ).count()

            week_start = today - timedelta(days=today.weekday())
            week_visits = VisitRecord.query.filter(
                func.date(VisitRecord.entry_time) >= week_start
            ).count()

            month_start = today.replace(day=1)
            month_visits = VisitRecord.query.filter(
                func.date(VisitRecord.entry_time) >= month_start
            ).count()

            year_start = today.replace(month=1, day=1)
            year_visits = VisitRecord.query.filter(
                func.date(VisitRecord.entry_time) >= year_start
            ).count()

            # 按状态统计申请 - 访问申请
            application_stats = db.session.query(
                VisitApplication.application_status,
                func.count(VisitApplication.id).label('count')
            ).group_by(VisitApplication.application_status).all()

            # 转换为字典格式
            application_stats = [{'status': stat.application_status, 'count': stat.count} for stat in application_stats]

            # 按学部统计校友
            division_stats = db.session.query(
                AlumniProfile.division,
                func.count(AlumniProfile.id).label('count')
            ).join(User, AlumniProfile.user_id == User.id).filter(
                User.user_type == 'alumni',
                AlumniProfile.approval_status == 'approved'
            ).group_by(AlumniProfile.division).order_by(
                func.count(AlumniProfile.id).desc()
            ).limit(10).all()

            # 按年级统计校友
            grade_stats = db.session.query(
                AlumniProfile.graduation_year,
                func.count(AlumniProfile.id).label('count')
            ).join(User, AlumniProfile.user_id == User.id).filter(
                User.user_type == 'alumni',
                AlumniProfile.approval_status == 'approved'
            ).group_by(AlumniProfile.graduation_year).order_by(
                AlumniProfile.graduation_year.desc()
            ).all()

            return jsonify({
                'time_statistics': {
                    'today_visits': today_visits,
                    'week_visits': week_visits,
                    'month_visits': month_visits,
                    'year_visits': year_visits,
                    'total_visits': total_visits
                },
                'application_stats': [{'status': stat.application_status, 'count': stat.count} for stat in application_stats],
                'division_stats': [{'division': stat.division, 'count': stat.count} for stat in division_stats],
                'grade_stats': [{'year': stat.graduation_year, 'count': stat.count} for stat in grade_stats],
                'total_applications': total_applications,
                'total_users': total_users,
                'total_alumni': total_alumni
            }), 200

        elif statistic_type == 'visits':
            # 访问详细统计
            # 按受访人统计
            target_person_stats = db.session.query(
                VisitApplication.target_person,
                func.count(VisitApplication.id).label('count')
            ).filter(
                VisitApplication.application_status.in_(['approved', 'completed'])
            ).group_by(VisitApplication.target_person).order_by(
                func.count(VisitApplication.id).desc()
            ).limit(20).all()

            # 按访问人统计
            visitor_stats = db.session.query(
                User.real_name,
                func.count(VisitApplication.id).label('count')
            ).join(VisitApplication).filter(
                VisitApplication.application_status.in_(['approved', 'completed'])
            ).group_by(User.real_name).order_by(
                func.count(VisitApplication.id).desc()
            ).limit(20).all()

            # 按审批人统计
            approver_stats = db.session.query(
                User.real_name,
                func.count(VisitApplication.id).label('count')
            ).join(VisitApplication, User.id == VisitApplication.approved_by).filter(
                VisitApplication.application_status.in_(['approved', 'completed'])
            ).group_by(User.real_name).order_by(
                func.count(VisitApplication.id).desc()
            ).limit(20).all()

            # 按放行人统计（保安）
            security_stats = db.session.query(
                User.real_name,
                func.count(VisitRecord.id).label('count')
            ).join(VisitRecord).filter(
                User.user_type == 'security'
            ).group_by(User.real_name).order_by(
                func.count(VisitRecord.id).desc()
            ).limit(20).all()

            # 按访问目的统计
            purpose_stats = db.session.query(
                VisitApplication.visit_purpose,
                func.count(VisitApplication.id).label('count')
            ).filter(
                VisitApplication.application_status.in_(['approved', 'completed'])
            ).group_by(VisitApplication.visit_purpose).order_by(
                func.count(VisitApplication.id).desc()
            ).all()

            return jsonify({
                'target_person_stats': [{'name': stat.target_person, 'count': stat.count} for stat in target_person_stats],
                'visitor_stats': [{'name': stat.real_name, 'count': stat.count} for stat in visitor_stats],
                'approver_stats': [{'name': stat.real_name, 'count': stat.count} for stat in approver_stats],
                'security_stats': [{'name': stat.real_name, 'count': stat.count} for stat in security_stats],
                'purpose_stats': [{'purpose': stat.visit_purpose, 'count': stat.count} for stat in purpose_stats]
            }), 200

        elif statistic_type == 'people':
            # 人员统计
            # 按班级统计校友
            class_stats = db.session.query(
                AlumniProfile.class_name,
                func.count(AlumniProfile.id).label('count')
            ).join(User, AlumniProfile.user_id == User.id).filter(
                User.user_type == 'alumni',
                AlumniProfile.approval_status == 'approved'
            ).group_by(AlumniProfile.class_name).order_by(
                func.count(AlumniProfile.id).desc()
            ).limit(20).all()

            # 按年级和班级统计
            grade_class_stats = db.session.query(
                AlumniProfile.graduation_year,
                AlumniProfile.class_name,
                func.count(AlumniProfile.id).label('count')
            ).join(User, AlumniProfile.user_id == User.id).filter(
                User.user_type == 'alumni',
                AlumniProfile.approval_status == 'approved'
            ).group_by(
                AlumniProfile.graduation_year,
                AlumniProfile.class_name
            ).order_by(
                AlumniProfile.graduation_year.desc(),
                func.count(AlumniProfile.id).desc()
            ).limit(30).all()

            return jsonify({
                'class_stats': [{'class_name': stat.class_name, 'count': stat.count} for stat in class_stats],
                'grade_class_stats': [
                    {
                        'graduation_year': stat.graduation_year,
                        'class_name': stat.class_name,
                        'count': stat.count
                    } for stat in grade_class_stats
                ]
            }), 200

        
        else:
            return jsonify({'error': '统计类型无效'}), 400

    except Exception as e:
        current_app.logger.error(f"获取统计数据失败: {str(e)}")
        return jsonify({'error': '获取统计数据失败', 'details': str(e)}), 500

@admin_bp.route('/users/<int:user_id>', methods=['GET'])
def get_user_detail(user_id):
    """获取用户详细信息"""
    try:
        current_app.logger.info(f"开始获取用户详情，用户ID: {user_id}")
        user = User.query.get(user_id)
        if not user:
            current_app.logger.warning(f"用户不存在，ID: {user_id}")
            return jsonify({'error': '用户不存在'}), 404

        current_app.logger.info(f"找到用户: {user.username}")
        user_data = user.to_dict()

        # 安全地添加组织名称，避免关系访问错误
        try:
            if hasattr(user, 'organization') and user.organization:
                user_data['organization_name'] = user.organization.name
        except Exception as org_error:
            current_app.logger.warning(f"获取组织信息失败: {org_error}")
            user_data['organization_name'] = None

        current_app.logger.info(f"用户数据转换成功")
        return jsonify({
            'success': True,
            'user': user_data
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取用户详情失败: {str(e)}")
        import traceback
        current_app.logger.error(f"错误堆栈: {traceback.format_exc()}")
        return jsonify({'error': '获取用户详情失败'}), 500

@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """更新用户信息"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': '用户不存在'}), 404

        data = request.get_json()

        # 更新基本信息
        if 'real_name' in data:
            user.real_name = data['real_name']
        if 'email' in data:
            # 检查邮箱是否已被其他用户使用
            existing_user = User.query.filter(User.email == data['email'], User.id != user_id).first()
            if existing_user:
                return jsonify({'error': '邮箱已被其他用户使用'}), 400
            user.email = data['email']
        if 'phone' in data:
            user.phone = data['phone']
        if 'user_type' in data:
            user.user_type = data['user_type']
        if 'status' in data:
            user.status = data['status']
        if 'organization_id' in data:
            user.organization_id = data['organization_id']
        if 'student_id' in data:
            user.student_id = data['student_id']
        if 'employee_id' in data:
            # 只有当员工编号不为空时才检查重复
            if data['employee_id'] and data['employee_id'].strip():
                # 检查员工编号是否已被其他用户使用
                existing_user = User.query.filter(
                    User.employee_id == data['employee_id'],
                    User.id != user_id
                ).first()
                if existing_user:
                    return jsonify({'error': '员工编号已被其他用户使用'}), 400
            user.employee_id = data['employee_id'] if data['employee_id'] and data['employee_id'].strip() else None

        # 更新密码（如果提供）
        if 'password' in data and data['password']:
            from werkzeug.security import generate_password_hash
            user.password_hash = generate_password_hash(data['password'])

        # 设置可拜访权限
        if 'is_visitable' in data:
            # 允许所有用户手动设置可拜访权限
            user.is_visitable = data['is_visitable']
        else:
            # 如果没有手动设置，教师默认为可拜访
            if 'teacher' in str(user.user_type):
                user.is_visitable = True

        user.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '用户信息更新成功',
            'user': user.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新用户信息失败: {str(e)}")
        return jsonify({'error': '更新用户信息失败'}), 500

@admin_bp.route('/users/<int:user_id>/visitable', methods=['PUT'])
def update_user_visitable(user_id):
    """更新用户可拜访权限"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': '用户不存在'}), 404

        if user.user_type != 'teacher':
            return jsonify({'error': '只有教师用户可以设置可拜访权限'}), 400

        data = request.get_json()
        is_visitable = data.get('is_visitable', False)

        user.is_visitable = is_visitable
        user.updated_at = datetime.utcnow()
        db.session.commit()

        status_text = '可拜访' if is_visitable else '不可拜访'
        return jsonify({
            'success': True,
            'message': f'用户已设置为{status_text}',
            'user': user.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新用户可拜访权限失败: {str(e)}")
        return jsonify({'error': '更新用户可拜访权限失败'}), 500

@admin_bp.route('/student-exit-statistics', methods=['GET'])
def get_student_exit_statistics():
    """获取学生出校详细统计报告"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # 如果没有指定日期范围，默认为最近30天
        if not start_date:
            start_date = (datetime.utcnow() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        if not end_date:
            end_date = datetime.utcnow().date()
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # 基础统计数据（暂时移除日期过滤进行测试）
        total_applications = StudentExitApplication.query.count()

        approved_applications = StudentExitApplication.query.filter_by(
            application_status='approved'
        ).count()

        rejected_applications = StudentExitApplication.query.filter_by(
            application_status='rejected'
        ).count()

        pending_applications = StudentExitApplication.query.filter(
            StudentExitApplication.application_status.in_(['pending', 'teacher_approved', 'parent_approved'])
        ).count()

        # 审批效率统计（暂时移除日期过滤进行测试）
        parent_approved = StudentExitApplication.query.filter_by(
            parent_approval_status='approved'
        ).count()

        parent_rejected = StudentExitApplication.query.filter_by(
            parent_approval_status='rejected'
        ).count()

        teacher_approved = StudentExitApplication.query.filter_by(
            teacher_approval_status='approved'
        ).count()

        teacher_rejected = StudentExitApplication.query.filter_by(
            teacher_approval_status='rejected'
        ).count()

        # 出校原因统计
        reason_stats = db.session.query(
            StudentExitApplication.exit_reason,
            func.count(StudentExitApplication.id).label('count')
        ).filter(
            func.date(StudentExitApplication.created_at) >= start_date,
            func.date(StudentExitApplication.created_at) <= end_date,
            StudentExitApplication.application_status.in_(['approved', 'completed'])
        ).group_by(StudentExitApplication.exit_reason).order_by(func.count(StudentExitApplication.id).desc()).all()

        # 目的地统计
        destination_stats = db.session.query(
            StudentExitApplication.destination,
            func.count(StudentExitApplication.id).label('count')
        ).filter(
            func.date(StudentExitApplication.created_at) >= start_date,
            func.date(StudentExitApplication.created_at) <= end_date,
            StudentExitApplication.application_status.in_(['approved', 'completed'])
        ).group_by(StudentExitApplication.destination).order_by(func.count(StudentExitApplication.id).desc()).limit(10).all()

        # 时间分布统计（按小时）
        hourly_stats = db.session.query(
            func.extract('hour', StudentExitApplication.created_at).label('hour'),
            func.count(StudentExitApplication.id).label('count')
        ).filter(
            func.date(StudentExitApplication.created_at) >= start_date,
            func.date(StudentExitApplication.created_at) <= end_date
        ).group_by(func.extract('hour', StudentExitApplication.created_at)).all()

        # 每日趋势数据
        daily_trend = []
        current_date = start_date
        while current_date <= end_date:
            day_count = StudentExitApplication.query.filter(
                func.date(StudentExitApplication.created_at) == current_date
            ).count()
            daily_trend.append({
                'date': current_date.isoformat(),
                'count': day_count
            })
            current_date += timedelta(days=1)

        # 班级统计（如果有班级信息）
        class_stats = db.session.query(
            User.class_name,
            func.count(StudentExitApplication.id).label('count')
        ).join(
            StudentExitApplication, User.id == StudentExitApplication.student_id
        ).filter(
            func.date(StudentExitApplication.created_at) >= start_date,
            func.date(StudentExitApplication.created_at) <= end_date,
            User.class_name.isnot(None)
        ).group_by(User.class_name).order_by(func.count(StudentExitApplication.id).desc()).limit(10).all()

        # 审批时间分析（平均审批时长）
        approval_time_analysis = db.session.query(
            func.avg(
                func.extract('epoch', StudentExitApplication.updated_at - StudentExitApplication.created_at) / 3600
            ).label('avg_hours')
        ).filter(
            func.date(StudentExitApplication.created_at) >= start_date,
            func.date(StudentExitApplication.created_at) <= end_date,
            StudentExitApplication.application_status.in_(['approved', 'rejected'])
        ).first()

        avg_approval_hours = float(approval_time_analysis.avg_hours) if approval_time_analysis.avg_hours else 0

        return jsonify({
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days': (end_date - start_date).days + 1
            },
            'summary': {
                'total_applications': total_applications,
                'approved_applications': approved_applications,
                'rejected_applications': rejected_applications,
                'pending_applications': pending_applications,
                'approval_rate': round(approved_applications / total_applications * 100, 2) if total_applications > 0 else 0,
                'rejection_rate': round(rejected_applications / total_applications * 100, 2) if total_applications > 0 else 0
            },
            'approval_efficiency': {
                'parent_approved': parent_approved,
                'parent_rejected': parent_rejected,
                'teacher_approved': teacher_approved,
                'teacher_rejected': teacher_rejected,
                'parent_approval_rate': round(parent_approved / total_applications * 100, 2) if total_applications > 0 else 0,
                'teacher_approval_rate': round(teacher_approved / total_applications * 100, 2) if total_applications > 0 else 0
            },
            'exit_reasons': [{'reason': stat.exit_reason, 'count': stat.count} for stat in reason_stats],
            'destinations': [{'destination': stat.destination, 'count': stat.count} for stat in destination_stats],
            'hourly_distribution': [{'hour': int(stat.hour), 'count': stat.count} for stat in hourly_stats],
            'daily_trend': daily_trend,
            'class_distribution': [{'class_name': stat.class_name, 'count': stat.count} for stat in class_stats],
            'performance_metrics': {
                'avg_approval_time_hours': round(avg_approval_hours, 2),
                'applications_per_day': round(total_applications / ((end_date - start_date).days + 1), 2)
            }
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取学生出校统计失败: {str(e)}")
        return jsonify({'error': '获取学生出校统计失败', 'details': str(e)}), 500


@admin_bp.route('/users/batch-delete', methods=['DELETE'])
def batch_delete_users():
    """批量删除用户"""
    try:
        data = request.get_json()
        user_ids = data.get('user_ids', [])

        if not user_ids:
            return jsonify({'error': '请提供要删除的用户ID列表'}), 400

        # 验证用户ID列表
        if not isinstance(user_ids, list):
            return jsonify({'error': '用户ID列表格式错误'}), 400

        # 获取当前管理员用户
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)

        # 不能删除自己
        if current_user_id in user_ids:
            return jsonify({'error': '不能删除当前登录的管理员账户'}), 400

        # 查询要删除的用户
        users_to_delete = User.query.filter(User.id.in_(user_ids)).all()
        found_user_ids = [user.id for user in users_to_delete]

        # 检查是否有用户不存在
        missing_user_ids = set(user_ids) - set(found_user_ids)
        if missing_user_ids:
            return jsonify({
                'error': f'以下用户不存在: {list(missing_user_ids)}'
            }), 404

        # 检查是否有其他管理员用户
        admin_users = [user for user in users_to_delete if user.user_type == 'admin']
        if admin_users:
            return jsonify({
                'error': '不能删除其他管理员用户'
            }), 400

        # 软删除：将用户状态设为 inactive，而不是物理删除
        deleted_count = 0
        for user in users_to_delete:
            # 删除用户的相关数据（如果需要的话）
            # 这里可以添加删除校友档案、访问记录等的逻辑

            # 软删除用户
            user.status = 'inactive'
            user.updated_at = datetime.utcnow()
            user.username = f"deleted_{user.id}_{user.username}"  # 避免用户名冲突
            user.email = f"deleted_{user.id}_{user.email}"
            deleted_count += 1

        db.session.commit()

        current_app.logger.info(f"管理员 {current_user.username} 批量删除了 {deleted_count} 个用户")

        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 个用户',
            'deleted_count': deleted_count
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量删除用户失败: {str(e)}")
        return jsonify({'error': '批量删除用户失败', 'details': str(e)}), 500


@admin_bp.route('/users/import-preview', methods=['POST'])
def import_users_preview():
    """预览批量导入用户数据"""
    try:
        from werkzeug.utils import secure_filename
        import os
        import pandas as pd
        from io import BytesIO

        # 检查文件
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400

        # 检查文件扩展名
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': '请上传Excel文件 (.xlsx 或 .xls)'}), 400

        # 读取Excel文件
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 400

        # 检查必需的列
        required_columns = ['用户名', '真实姓名', '用户类型', '邮箱']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'error': f'缺少必需的列: {", ".join(missing_columns)}',
                'required_columns': required_columns
            }), 400

        # 验证数据
        users = []
        errors = []
        user_type_mapping = {
            '教师': 'teacher',
            '学生': 'student',
            '家长': 'parent',
            '校友': 'alumni',
            '保安': 'security',
            '管理员': 'admin'
        }

        for index, row in df.iterrows():
            user_data = {
                'username': str(row.get('用户名', '')).strip(),
                'real_name': str(row.get('真实姓名', '')).strip(),
                'user_type': user_type_mapping.get(str(row.get('用户类型', '')), 'teacher'),
                'email': str(row.get('邮箱', '')).strip(),
                'phone': str(row.get('手机号', '')).strip() if pd.notna(row.get('手机号')) else '',
                'student_id': str(row.get('学号', '')).strip() if pd.notna(row.get('学号')) else '',
                'employee_id': str(row.get('工号', '')).strip() if pd.notna(row.get('工号')) else '',
                'class_id': str(row.get('班级', '')).strip() if pd.notna(row.get('班级')) else '',
                'grade': str(row.get('年级', '')).strip() if pd.notna(row.get('年级')) else ''
            }

            # 验证必填字段
            is_valid = True
            validation_errors = []

            if not user_data['username']:
                validation_errors.append('用户名不能为空')
                is_valid = False

            if not user_data['real_name']:
                validation_errors.append('真实姓名不能为空')
                is_valid = False

            if not user_data['email']:
                validation_errors.append('邮箱不能为空')
                is_valid = False

            # 验证邮箱格式
            import re
            email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if user_data['email'] and not re.match(email_pattern, user_data['email']):
                validation_errors.append('邮箱格式不正确')
                is_valid = False

            user_data['valid'] = is_valid
            user_data['validation_errors'] = validation_errors
            users.append(user_data)

        return jsonify({
            'success': True,
            'data': {
                'users': users,
                'total': len(users),
                'valid_count': sum(1 for u in users if u['valid']),
                'invalid_count': sum(1 for u in users if not u['valid'])
            }
        })

    except Exception as e:
        current_app.logger.error(f"预览导入用户数据失败: {str(e)}")
        return jsonify({'error': '预览导入失败', 'details': str(e)}), 500


@admin_bp.route('/users/import', methods=['POST'])
def import_users():
    """批量导入用户"""
    try:
        data = request.get_json()
        users = data.get('users', [])

        if not users:
            return jsonify({'error': '没有用户数据'}), 400

        success_count = 0
        failed_count = 0
        errors = []
        warnings = []

        for user_data in users:
            try:
                # 检查用户名是否已存在
                existing_user = User.query.filter_by(username=user_data['username']).first()
                if existing_user:
                    warnings.append(f"用户名 '{user_data['username']}' 已存在，跳过")
                    continue

                # 检查邮箱是否已存在
                existing_email = User.query.filter_by(email=user_data['email']).first()
                if existing_email:
                    warnings.append(f"邮箱 '{user_data['email']}' 已存在，跳过")
                    continue

                # 创建新用户
                new_user = User(
                    username=user_data['username'],
                    real_name=user_data['real_name'],
                    user_type=user_data['user_type'],
                    email=user_data['email'],
                    phone=user_data.get('phone') or None,
                    student_id=user_data.get('student_id') or None,
                    employee_id=user_data.get('employee_id') or None,
                    class_id=user_data.get('class_id') or None,
                    grade=user_data.get('grade') or None,
                    status='active',
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow()
                )

                # 设置默认密码（用户名+123456）
                default_password = user_data['username'] + '123456'
                new_user.set_password(default_password)

                db.session.add(new_user)
                db.session.commit()
                success_count += 1

            except Exception as e:
                failed_count += 1
                errors.append(f"用户 '{user_data.get('username', 'unknown')}' 创建失败: {str(e)}")
                db.session.rollback()

        return jsonify({
            'success': True,
            'data': {
                'total': len(users),
                'success': success_count,
                'failed': failed_count,
                'warnings': len(warnings),
                'errors': errors,
                'warning_details': warnings
            }
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量导入用户失败: {str(e)}")
        return jsonify({'error': '批量导入失败', 'details': str(e)}), 500


@admin_bp.route('/users/template', methods=['GET'])
def download_user_template():
    """下载用户导入模板"""
    try:
        import pandas as pd
        from io import BytesIO

        # 创建模板数据
        template_data = {
            '用户名*': ['zhangsan', 'lisi', 'wangwu'],
            '密码*': ['password123', 'password123', 'password123'],
            '真实姓名*': ['张三', '李四', '王五'],
            '邮箱*': ['zhangsan@example.com', 'lisi@example.com', 'wangwu@example.com'],
            '手机号': ['13800138000', '13800138001', '13800138002'],
            '用户类型*': ['teacher', 'student', 'alumni'],
            '学号': ['', 'S2023001', 'S2020001'],
            '工号': ['T001', '', ''],
            '班级': ['高三1班', '高三1班', ''],
            '年级': ['高三', '高三', ''],
            '是否班主任': ['是', '', ''],
            '可拜访权限': ['是', '', '是']
        }

        df = pd.DataFrame(template_data)

        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='用户导入模板', index=False)

        output.seek(0)

        return current_app.response_class(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=用户导入模板.xlsx'}
        )

    except Exception as e:
        current_app.logger.error(f"下载用户模板失败: {str(e)}")
        return jsonify({'error': '下载模板失败', 'details': str(e)}), 500


@admin_bp.route('/users/export', methods=['GET'])
def export_users():
    """导出用户数据"""
    try:
        import pandas as pd
        from io import BytesIO

        # 获取查询参数
        search = request.args.get('search', '')
        user_type = request.args.get('user_type', '')
        status = request.args.get('status', '')

        # 构建查询
        query = User.query
        if search:
            query = query.filter(
                User.username.contains(search) |
                User.real_name.contains(search) |
                User.email.contains(search)
            )
        if user_type:
            query = query.filter(User.user_type == user_type)
        if status:
            query = query.filter(User.status == status)

        users = query.all()

        # 转换为DataFrame
        data = []
        for user in users:
            data.append({
                '用户名': user.username,
                '真实姓名': user.real_name,
                '用户类型': user.user_type,
                '邮箱': user.email,
                '手机号': user.phone or '',
                '状态': user.status,
                '学号': user.student_id or '',
                '工号': user.employee_id or '',
                '班级': user.class_id or '',
                '年级': user.grade or '',
                '创建时间': user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
                '最后更新': user.updated_at.strftime('%Y-%m-%d %H:%M:%S') if user.updated_at else ''
            })

        df = pd.DataFrame(data)

        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='用户数据', index=False)

        output.seek(0)

        return current_app.response_class(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=用户数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
        )

    except Exception as e:
        current_app.logger.error(f"导出用户数据失败: {str(e)}")
        return jsonify({'error': '导出失败', 'details': str(e)}), 500


# ===================== 组织管理API =====================

@admin_bp.route('/organizations', methods=['GET'])
def get_organizations():
    """获取组织列表"""
    try:
        # 获取查询参数
        search = request.args.get('search', '')
        org_type = request.args.get('org_type', '')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        # 构建查询
        query = Organization.query
        if search:
            query = query.filter(
                Organization.name.contains(search) |
                Organization.code.contains(search) |
                Organization.description.contains(search)
            )
        if org_type:
            query = query.filter(Organization.org_type == org_type)

        # 按层级和排序排列
        query = query.order_by(Organization.level, Organization.sort_order, Organization.name)

        # 分页
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        organizations = pagination.items

        return jsonify({
            'organizations': [org.to_dict(include_children=False, include_users=True) for org in organizations],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        })

    except Exception as e:
        current_app.logger.error(f"获取组织列表失败: {str(e)}")
        return jsonify({'error': '获取组织列表失败', 'details': str(e)}), 500


@admin_bp.route('/organizations/tree', methods=['GET'])
def get_organization_tree():
    """获取组织树结构"""
    try:
        # 获取所有活跃组织
        organizations = Organization.query.filter_by(status='active').order_by(Organization.level, Organization.sort_order, Organization.name).all()

        # 构建树结构
        def build_tree(orgs, parent_id=None):
            tree = []
            for org in orgs:
                if org.parent_id == parent_id:
                    node = org.to_dict(include_children=False, include_users=True)
                    children = build_tree(orgs, org.id)
                    if children:
                        node['children'] = children
                    tree.append(node)
            return tree

        tree = build_tree(organizations)

        return jsonify({
            'tree': tree,
            'total': len(organizations)
        })

    except Exception as e:
        current_app.logger.error(f"获取组织树失败: {str(e)}")
        return jsonify({'error': '获取组织树失败', 'details': str(e)}), 500


@admin_bp.route('/organizations', methods=['POST'])
def create_organization():
    """创建组织"""
    try:
        data = request.get_json()

        # 验证必需字段
        required_fields = ['name', 'code', 'org_type']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'缺少必需字段: {field}'}), 400

        # 检查代码是否重复
        existing_org = Organization.query.filter_by(code=data['code']).first()
        if existing_org:
            return jsonify({'error': '组织代码已存在'}), 400

        # 创建组织
        organization = Organization(
            name=data['name'],
            code=data['code'],
            org_type=data['org_type'],
            description=data.get('description', ''),
            parent_id=data.get('parent_id'),
            contact_person=data.get('contact_person'),
            contact_phone=data.get('contact_phone'),
            contact_email=data.get('contact_email'),
            address=data.get('address'),
            status=data.get('status', 'active'),
            sort_order=data.get('sort_order', 0),
            created_by=get_jwt_identity()
        )

        # 设置路径和层级
        organization.before_save()

        db.session.add(organization)
        db.session.commit()

        return jsonify({
            'message': '组织创建成功',
            'organization': organization.to_dict(include_children=False, include_users=True)
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"创建组织失败: {str(e)}")
        return jsonify({'error': '创建组织失败', 'details': str(e)}), 500


@admin_bp.route('/organizations/<int:org_id>', methods=['GET'])
def get_organization(org_id):
    """获取组织详情"""
    try:
        organization = Organization.query.get_or_404(org_id)

        return jsonify({
            'organization': organization.to_dict(include_children=True, include_users=True)
        })

    except Exception as e:
        current_app.logger.error(f"获取组织详情失败: {str(e)}")
        return jsonify({'error': '获取组织详情失败', 'details': str(e)}), 500


@admin_bp.route('/organizations/<int:org_id>', methods=['PUT'])
def update_organization(org_id):
    """更新组织"""
    try:
        organization = Organization.query.get_or_404(org_id)
        data = request.get_json()

        # 更新基本信息
        if 'name' in data:
            organization.name = data['name']
        if 'description' in data:
            organization.description = data['description']
        if 'contact_person' in data:
            organization.contact_person = data['contact_person']
        if 'contact_phone' in data:
            organization.contact_phone = data['contact_phone']
        if 'contact_email' in data:
            organization.contact_email = data['contact_email']
        if 'address' in data:
            organization.address = data['address']
        if 'status' in data:
            organization.status = data['status']
        if 'sort_order' in data:
            organization.sort_order = data['sort_order']

        # 更新关系字段
        if 'class_teacher_id' in data:
            organization.class_teacher_id = data['class_teacher_id']
        if 'head_teacher_id' in data:
            organization.head_teacher_id = data['head_teacher_id']
        if 'leader_id' in data:
            organization.leader_id = data['leader_id']

        # 如果更改了父组织，更新路径和层级
        if 'parent_id' in data and organization.parent_id != data['parent_id']:
            organization.parent_id = data['parent_id']
            organization.before_save()

        organization.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'message': '组织更新成功',
            'organization': organization.to_dict(include_children=False, include_users=True)
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新组织失败: {str(e)}")
        return jsonify({'error': '更新组织失败', 'details': str(e)}), 500


@admin_bp.route('/organizations/<int:org_id>', methods=['DELETE'])
def delete_organization(org_id):
    """删除组织"""
    try:
        organization = Organization.query.get_or_404(org_id)

        # 检查是否有子组织
        children = Organization.query.filter_by(parent_id=org_id).all()
        if children:
            return jsonify({'error': '无法删除包含子组织的组织'}), 400

        # 检查是否有关联用户
        users = User.query.filter_by(organization_id=org_id).all()
        if users:
            return jsonify({'error': '无法删除包含用户的组织'}), 400

        db.session.delete(organization)
        db.session.commit()

        return jsonify({'message': '组织删除成功'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除组织失败: {str(e)}")
        return jsonify({'error': '删除组织失败', 'details': str(e)}), 500


@admin_bp.route('/organizations/<int:org_id>/available-teachers', methods=['GET'])
def get_available_teachers(org_id):
    """获取可选择的教师列表"""
    try:
        # 获取当前组织
        organization = Organization.query.get_or_404(org_id)

        # 获取所有教师用户
        teachers = User.query.filter_by(user_type='teacher', status='active').all()

        # 构建教师列表
        teacher_list = []
        for teacher in teachers:
            teacher_info = teacher.to_dict()
            # 标记当前已选中的教师
            teacher_info['is_selected'] = {
                'class_teacher': teacher.id == organization.class_teacher_id,
                'head_teacher': teacher.id == organization.head_teacher_id,
                'leader': teacher.id == organization.leader_id
            }
            teacher_list.append(teacher_info)

        return jsonify({
            'teachers': teacher_list,
            'current_assignments': {
                'class_teacher_id': organization.class_teacher_id,
                'head_teacher_id': organization.head_teacher_id,
                'leader_id': organization.leader_id
            }
        })

    except Exception as e:
        current_app.logger.error(f"获取教师列表失败: {str(e)}")
        return jsonify({'error': '获取教师列表失败', 'details': str(e)}), 500


@admin_bp.route('/users/<int:user_id>/relationships', methods=['GET'])
def get_user_relationships(user_id):
    """获取用户关系信息"""
    try:
        user = User.query.get_or_404(user_id)

        # 获取详细关系信息
        user_data = user.to_dict(include_sensitive=True)

        # 获取可关联的用户列表
        potential_relationships = {}

        if user.user_type == 'parent':
            # 获取可关联的学生
            available_students = User.query.filter_by(user_type='student', status='active').all()
            potential_relationships['students'] = [
                {
                    'id': student.id,
                    'name': student.real_name,
                    'student_id': student.student_id,
                    'is_related': student.id in [child.id for child in user.parent_students] if hasattr(user, 'parent_students') else False
                }
                for student in available_students
            ]

        elif user.user_type == 'student':
            # 获取可关联的家长
            available_parents = User.query.filter_by(user_type='parent', status='active').all()
            potential_relationships['parents'] = [
                {
                    'id': parent.id,
                    'name': parent.real_name,
                    'phone': parent.phone,
                    'is_related': parent.id in [par.id for par in user.student_parents] if hasattr(user, 'student_parents') else False
                }
                for parent in available_parents
            ]

        return jsonify({
            'user': user_data,
            'potential_relationships': potential_relationships
        })

    except Exception as e:
        current_app.logger.error(f"获取用户关系失败: {str(e)}")
        return jsonify({'error': '获取用户关系失败', 'details': str(e)}), 500


@admin_bp.route('/users/<int:user_id>/relationships', methods=['PUT'])
def update_user_relationships(user_id):
    """更新用户关系"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()

        # 更新家长-学生关系
        if user.user_type == 'parent' and 'student_ids' in data:
            # 清除现有关系
            if hasattr(user, 'parent_students'):
                for student in user.parent_students:
                    student.student_parent_id = None

            # 建立新关系
            for student_id in data['student_ids']:
                student = User.query.get(student_id)
                if student and student.user_type == 'student':
                    student.student_parent_id = user.id

        elif user.user_type == 'student' and 'parent_ids' in data:
            # 清除现有关系
            if hasattr(user, 'student_parents'):
                for parent in user.student_parents:
                    parent.parent_student_id = None

            # 建立新关系
            for parent_id in data['parent_ids']:
                parent = User.query.get(parent_id)
                if parent and parent.user_type == 'parent':
                    parent.parent_student_id = user.id

        # 更新用户组织关系
        if 'organization_id' in data:
            user.organization_id = data['organization_id']

        user.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'message': '用户关系更新成功',
            'user': user.to_dict(include_sensitive=True)
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新用户关系失败: {str(e)}")
        return jsonify({'error': '更新用户关系失败', 'details': str(e)}), 500

@admin_bp.route('/export/card-data', methods=['GET'])
def export_card_data():
    """导出制卡中心数据"""
    try:
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        import os

        # 查询所有活跃用户（排除管理员）
        users = User.query.filter(
            User.user_type.in_(['alumni', 'student', 'teacher']),
            User.status == 'active'
        ).all()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "制卡数据导入"

        # 设置列标题（根据模板的11列）
        headers = [
            '学生姓名', '客户名称', '一卡通号', '身份证号', '民族',
            '文件名称', '卡类型（0-29）', '客户编号', '证件类型',
            '性别（1为男，0为女）', '手机号'
        ]

        # 设置标题样式
        header_font = Font(name='宋体', size=12, bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 写入标题行
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置数据样式
        data_font = Font(name='宋体', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center')

        # 写入用户数据
        for row_idx, user in enumerate(users, 2):
            # 根据用户类型设置客户名称
            if user.user_type == 'alumni':
                customer_name = f"校友-{user.real_name}"
                card_type = "1"  # 校友卡类型
                customer_no = user.student_id or f"A{user.id:06d}"
            elif user.user_type == 'student':
                customer_name = f"学生-{user.real_name}"
                card_type = "2"  # 学生卡类型
                customer_no = user.student_id or f"S{user.id:06d}"
            elif user.user_type == 'teacher':
                customer_name = f"教师-{user.real_name}"
                card_type = "3"  # 教师卡类型
                customer_no = user.employee_id or f"T{user.id:06d}"
            else:
                customer_name = user.real_name
                card_type = "0"
                customer_no = f"U{user.id:06d}"

            # 性别转换（1为男，0为女）
            gender_code = "1" if user.gender == '男' else "0"

            # 证件类型（1为身份证，2为护照等）
            id_type = "1" if user.id_card and len(user.id_card) == 18 else "2"

            # 文件名称（学号/工号_姓名）
            file_name = f"{customer_no}_{user.real_name}"

            # 写入数据行
            data_row = [
                user.real_name,                    # 学生姓名
                customer_name,                     # 客户名称
                customer_no,                       # 一卡通号
                user.id_card or "",                # 身份证号
                "汉族",                            # 民族（默认汉族）
                file_name,                         # 文件名称
                card_type,                         # 卡类型
                customer_no,                       # 客户编号
                id_type,                           # 证件类型
                gender_code,                       # 性别
                user.phone or ""                   # 手机号
            ]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment

        # 调整列宽
        column_widths = [12, 20, 15, 20, 8, 20, 15, 15, 10, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 保存到内存中的文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"制卡中心数据导出_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        current_app.logger.error(f"导出制卡数据失败: {str(e)}")
        return jsonify({'error': '导出制卡数据失败', 'details': str(e)}), 500

